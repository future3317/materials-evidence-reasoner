#!/usr/bin/env python3
"""Create a conservative, loss-aware intake packet from materials R&D files.

The script inventories files, preserves raw values, maps common table headers to
canonical intake fields, and produces a prioritized missing-information report.
It deliberately does not infer scientific facts from PDFs/images or fill missing
process parameters with domain defaults.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import os
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

ROOT = Path(__file__).resolve().parents[1]
ALIASES_PATH = ROOT / "references" / "intake-field-aliases.json"
SCHEMA_PATH = ROOT / "references" / "input-schema.json"

TEXT_EXTENSIONS = {".txt", ".md", ".markdown", ".rst", ".log"}
PYTHON_EXTENSIONS = {".py"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp", ".gif"}
CSV_EXTENSIONS = {".csv", ".tsv"}
IGNORED_NAMES = {".ds_store", "thumbs.db"}


def is_metadata_file(path: Path) -> bool:
    """Ignore archive metadata that is not a scientific input artifact."""
    return path.name.lower() in IGNORED_NAMES or path.name.startswith("._") or any(
        part.casefold() == "__macosx" for part in path.parts
    )


def normalize_token(value: str) -> str:
    value = unicodedata.normalize("NFKC", value).strip().lower()
    value = re.sub(r"[\s_\-/]+", " ", value)
    value = value.strip(" .:：()[]{}")
    return value


def load_aliases() -> tuple[dict[str, str], set[str], dict[str, Any]]:
    spec = json.loads(ALIASES_PATH.read_text(encoding="utf-8"))
    index: dict[str, str] = {}
    canonical_names: set[str] = set(spec["fields"])
    for canonical, aliases in spec["fields"].items():
        index[normalize_token(canonical)] = canonical
        for alias in aliases:
            token = normalize_token(alias)
            existing = index.get(token)
            if existing is not None and existing != canonical:
                raise ValueError(f"alias collision: {alias!r} -> {existing!r}/{canonical!r}")
            index[token] = canonical
    return index, canonical_names, spec


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def relative_artifact_path(path: Path, base: Path) -> str:
    """Store input references relative to the intake packet directory."""
    try:
        return Path(os.path.relpath(path.resolve(), base.resolve())).as_posix()
    except ValueError:
        return path.name


def detect_kind(path: Path) -> tuple[str, list[str]]:
    ext = path.suffix.lower()
    warnings: list[str] = []
    try:
        head = path.read_bytes()[:16]
    except OSError as exc:
        return "binary", [f"cannot read file header: {exc}"]

    if ext == ".pdf":
        if not head.startswith(b"%PDF"):
            warnings.append("extension is .pdf but PDF magic bytes are missing")
        return "pdf", warnings
    if ext == ".xlsx":
        if not head.startswith(b"PK"):
            warnings.append("extension is .xlsx but ZIP magic bytes are missing")
        return "xlsx", warnings
    if ext in {".xml", ".nxml", ".jats", ".tei"}:
        if not head.lstrip().startswith(b"<"):
            warnings.append("extension is XML-like but an opening tag was not found")
        return "xml", warnings
    if ext in {".html", ".htm", ".xhtml"}:
        return "html", warnings
    if ext == ".docx":
        if not head.startswith(b"PK"):
            warnings.append("extension is .docx but ZIP magic bytes are missing")
        return "docx", warnings
    if ext == ".epub":
        if not head.startswith(b"PK"):
            warnings.append("extension is .epub but ZIP magic bytes are missing")
        return "epub", warnings
    if ext in CSV_EXTENSIONS:
        return "tsv" if ext == ".tsv" else "csv", warnings
    if ext == ".json":
        return "json", warnings
    if ext in TEXT_EXTENSIONS or ext in PYTHON_EXTENSIONS:
        return "markdown" if ext in {".md", ".markdown"} else "text", warnings
    if ext in IMAGE_EXTENSIONS:
        signatures = {
            ".png": b"\x89PNG",
            ".jpg": b"\xff\xd8\xff",
            ".jpeg": b"\xff\xd8\xff",
            ".gif": b"GIF8",
            ".tif": (b"II*\x00", b"MM\x00*"),
            ".tiff": (b"II*\x00", b"MM\x00*"),
        }
        expected = signatures.get(ext)
        if expected:
            ok = any(head.startswith(item) for item in expected) if isinstance(expected, tuple) else head.startswith(expected)
            if not ok:
                warnings.append(f"extension is {ext} but expected image magic bytes are missing")
        return "image", warnings
    return "binary", warnings


def iter_files(root_input: Path, output: Path | None = None) -> list[Path]:
    if root_input.is_file():
        return [root_input]
    if not root_input.is_dir():
        raise FileNotFoundError(f"input path does not exist: {root_input}")
    output_resolved = output.resolve() if output is not None else None
    return [
        p for p in sorted(root_input.rglob("*"))
        if p.is_file()
        and not is_metadata_file(p)
        and not (output_resolved is not None and (p.resolve() == output_resolved or output_resolved in p.resolve().parents))
    ]


def read_text_with_fallback(path: Path) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return path.read_text(encoding=encoding), encoding
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError("unknown", b"", 0, 1, "no supported encoding")


def map_header(raw_name: str, alias_index: dict[str, str], canonical_names: set[str], ambiguous: dict[str, str]) -> dict[str, Any]:
    token = normalize_token(raw_name)
    canonical = alias_index.get(token)
    if token in ambiguous:
        return {
            "raw_name": raw_name,
            "canonical_name": canonical,
            "mapping_status": "contextual-candidate",
            "mapping_note": ambiguous[token],
        }
    if canonical is None:
        return {
            "raw_name": raw_name,
            "canonical_name": None,
            "mapping_status": "unresolved",
            "mapping_note": None,
        }
    exact = token == normalize_token(canonical) and canonical in canonical_names
    return {
        "raw_name": raw_name,
        "canonical_name": canonical,
        "mapping_status": "exact" if exact else "alias",
        "mapping_note": None,
    }


def unique_headers(headers: Iterable[Any]) -> list[str]:
    output: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(headers, start=1):
        base = str(value).strip() if value is not None and str(value).strip() else f"unnamed_{index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        output.append(base if count == 1 else f"{base}__{count}")
    return output


def table_from_rows(
    input_id: str,
    table_id: str,
    source_name: str,
    rows: list[list[Any]],
    alias_index: dict[str, str],
    canonical_names: set[str],
    ambiguous: dict[str, str],
    sheet_name: str | None = None,
    warnings: list[str] | None = None,
) -> dict[str, Any] | None:
    if not rows:
        return None
    header_index = next((i for i, row in enumerate(rows) if any(cell not in (None, "") for cell in row)), None)
    if header_index is None:
        return None
    headers = unique_headers(rows[header_index])
    columns = [map_header(name, alias_index, canonical_names, ambiguous) for name in headers]
    data_rows: list[dict[str, Any]] = []
    for raw_row in rows[header_index + 1 :]:
        if not any(value not in (None, "") for value in raw_row):
            continue
        padded = list(raw_row) + [None] * max(0, len(headers) - len(raw_row))
        raw_record = {headers[i]: padded[i] for i in range(len(headers))}
        canonical_record: dict[str, Any] = {}
        for index, column in enumerate(columns):
            canonical = column["canonical_name"]
            if canonical and column["mapping_status"] in {"exact", "alias"}:
                value = padded[index]
                if canonical in canonical_record and canonical_record[canonical] not in (None, "", value):
                    canonical_record.setdefault(f"{canonical}__conflict", []).append(value)
                else:
                    canonical_record[canonical] = value
        data_rows.append({"raw": raw_record, "canonical": canonical_record})
    return {
        "input_id": input_id,
        "table_id": table_id,
        "source_name": source_name,
        "sheet_name": sheet_name,
        "row_count": len(data_rows),
        "columns": columns,
        "rows": data_rows,
        "warnings": warnings or [],
    }


def parse_delimited(path: Path, input_id: str, alias_index: dict[str, str], canonical_names: set[str], ambiguous: dict[str, str]) -> list[dict[str, Any]]:
    text, encoding = read_text_with_fallback(path)
    sample = text[:8192]
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        pass
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows = [row for row in reader]
    warnings = [f"decoded with {encoding}", f"delimiter={delimiter!r}"]
    table = table_from_rows(input_id, f"{input_id}-T1", path.name, rows, alias_index, canonical_names, ambiguous, warnings=warnings)
    return [table] if table else []


def parse_xlsx(path: Path, input_id: str, alias_index: dict[str, str], canonical_names: set[str], ambiguous: dict[str, str]) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        import openpyxl
    except ImportError:
        return [], ["openpyxl is unavailable; workbook recorded as manifest-only"]
    warnings: list[str] = []
    try:
        workbook_formula = openpyxl.load_workbook(path, data_only=False, read_only=False)
        workbook_values = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        return [], [f"workbook parse failed: {exc}"]
    tables: list[dict[str, Any]] = []
    for sheet_index, sheet in enumerate(workbook_formula.worksheets, start=1):
        value_sheet = workbook_values[sheet.title]
        rows: list[list[Any]] = []
        formula_count = 0
        missing_cache_count = 0
        for formula_row, value_row in zip(sheet.iter_rows(), value_sheet.iter_rows()):
            output_row: list[Any] = []
            for formula_cell, value_cell in zip(formula_row, value_row):
                value = formula_cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if value_cell.value is None:
                        missing_cache_count += 1
                    output_row.append({"formula": value, "cached_value": value_cell.value})
                else:
                    output_row.append(value)
            rows.append(output_row)
        sheet_warnings: list[str] = []
        if formula_count:
            sheet_warnings.append(f"contains {formula_count} formula cells; formulas were not executed")
        if missing_cache_count:
            sheet_warnings.append(f"{missing_cache_count} formula cells have no cached value")
        hidden_rows = sum(1 for dim in sheet.row_dimensions.values() if dim.hidden)
        hidden_cols = sum(1 for dim in sheet.column_dimensions.values() if dim.hidden)
        if hidden_rows or hidden_cols:
            sheet_warnings.append(f"hidden rows={hidden_rows}, hidden columns={hidden_cols}")
        if sheet.auto_filter and sheet.auto_filter.ref:
            sheet_warnings.append(f"worksheet has an autofilter: {sheet.auto_filter.ref}")
        merged = len(sheet.merged_cells.ranges)
        if merged:
            sheet_warnings.append(f"worksheet has {merged} merged ranges")
        table = table_from_rows(
            input_id,
            f"{input_id}-T{sheet_index}",
            path.name,
            rows,
            alias_index,
            canonical_names,
            ambiguous,
            sheet_name=sheet.title,
            warnings=sheet_warnings,
        )
        if table:
            tables.append(table)
    workbook_formula.close()
    workbook_values.close()
    return tables, warnings


def parse_json_file(path: Path, input_id: str, alias_index: dict[str, str], canonical_names: set[str], ambiguous: dict[str, str]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [], [], [f"JSON parse failed: {exc}"]
    tables: list[dict[str, Any]] = []
    text_items: list[dict[str, Any]] = []

    def rows_from_objects(items: list[dict[str, Any]], table_id: str, source_name: str) -> None:
        headers: list[str] = []
        for item in items:
            for key in item:
                if key not in headers:
                    headers.append(key)
        rows = [headers] + [[item.get(key) for key in headers] for item in items]
        table = table_from_rows(input_id, table_id, source_name, rows, alias_index, canonical_names, ambiguous)
        if table:
            tables.append(table)

    if isinstance(data, list) and all(isinstance(item, dict) for item in data):
        rows_from_objects(data, f"{input_id}-T1", path.name)
    elif isinstance(data, dict):
        list_tables = [(key, value) for key, value in data.items() if isinstance(value, list) and value and all(isinstance(item, dict) for item in value)]
        if list_tables:
            for index, (key, value) in enumerate(list_tables, start=1):
                rows_from_objects(value, f"{input_id}-T{index}", f"{path.name}#{key}")
        else:
            scalar = {key: value for key, value in data.items() if not isinstance(value, (dict, list))}
            if scalar:
                rows_from_objects([scalar], f"{input_id}-T1", path.name)
            text_items.append({"input_id": input_id, "text": json.dumps(data, ensure_ascii=False, indent=2), "extracted_pairs": []})
    else:
        text_items.append({"input_id": input_id, "text": json.dumps(data, ensure_ascii=False), "extracted_pairs": []})
    return tables, text_items, []


def parse_text(path: Path, input_id: str, alias_index: dict[str, str], canonical_names: set[str], ambiguous: dict[str, str]) -> tuple[dict[str, Any], list[str]]:
    try:
        text, encoding = read_text_with_fallback(path)
    except Exception as exc:
        return {"input_id": input_id, "text": "", "extracted_pairs": []}, [f"text parse failed: {exc}"]
    pairs: list[dict[str, Any]] = []
    for line in text.splitlines():
        match = re.match(r"^\s*[-*]?\s*([^:：]{1,80})\s*[:：]\s*(.+?)\s*$", line)
        if not match:
            continue
        raw_key, raw_value = match.groups()
        mapping = map_header(raw_key, alias_index, canonical_names, ambiguous)
        pairs.append({
            "raw_key": raw_key.strip(),
            "raw_value": raw_value.strip(),
            "canonical_name": mapping["canonical_name"],
            "mapping_status": mapping["mapping_status"],
        })
    return {"input_id": input_id, "text": text, "extracted_pairs": pairs}, [f"decoded with {encoding}"]


def canonical_fields(packet: dict[str, Any]) -> set[str]:
    fields: set[str] = set()
    for table in packet["normalized_tables"]:
        for column in table["columns"]:
            if column["canonical_name"] and column["mapping_status"] in {"exact", "alias"}:
                fields.add(column["canonical_name"])
        for row in table["rows"]:
            fields.update(row.get("canonical", {}).keys())
    for item in packet["text_items"]:
        for pair in item["extracted_pairs"]:
            if pair["canonical_name"] and pair["mapping_status"] in {"exact", "alias"}:
                fields.add(pair["canonical_name"])
    return {field.split("__conflict", 1)[0] for field in fields}


def build_missing(packet: dict[str, Any], decision_question: str | None) -> list[dict[str, str]]:
    fields = canonical_fields(packet)
    has_tables = bool(packet["normalized_tables"])
    missing: list[dict[str, str]] = []

    def add(field: str, priority: str, reason: str, impact: str, suggested: str) -> None:
        missing.append({"field": field, "priority": priority, "reason": reason, "impact": impact, "suggested_input": suggested})

    if not decision_question:
        add("decision_question", "analysis-limiting", "未提供明确的研发决策问题。", "Agent 可能围绕错误目标分析，或输出过于宽泛的报告。", "用一句话说明要做的决策，例如“判断该批次是否偏离”或“区分哪两个机理”。")
    if not ({"material_name", "composition"} & fields):
        add("material_identity", "blocker" if has_tables else "analysis-limiting", "未识别到材料名称或组成。", "性能值无法可靠绑定到正确的材料体系。", "提供材料名称或化学式，并说明适用于哪些行或样品。")
    if not ({"sample_id", "batch_id"} & fields):
        add("sample_or_batch_id", "analysis-limiting", "未识别到样品或批次标识。", "不同试样或批次可能被错误合并，也无法评估重复性。", "增加稳定的 sample_id 或 batch_id 列。")
    if "property_value" in fields and "unit" not in fields:
        add("unit", "blocker", "发现数值结果，但未识别到单位。", "单位解释不同可能改变物理意义并使比较失效。", "在独立列或明确表头中补充单位，并说明归一化分母。")
    if "property_value" not in fields:
        add("property_value", "analysis-limiting", "未识别到测量性能值。", "无法执行定量偏差检测。", "提供目标测量值，或上传仪器导出表。")
    if "property_name" not in fields:
        add("property_name", "analysis-limiting", "未识别到性能或指标名称。", "不知道测量对象时，数值无法被正确解释。", "为每个数值列标注性能或指标名称。")
    if "measurement_method" not in fields:
        add("measurement_method", "analysis-limiting", "缺少测试或表征方法。", "依赖方法的数值可能被错误比较。", "提供方法、标准或仪器，以及关键测试协议。")
    if "normalization_basis" not in fields and "property_value" in fields:
        add("normalization_basis", "analysis-limiting", "未识别到分母或归一化基准。", "按质量、面积、体积或活性物质归一化的数值可能不可比。", "说明分母、几何尺寸、活性组分比例或计算公式。")
    if not ({"replicate_count", "uncertainty", "measurement_repeat_id"} & fields):
        add("replicates_and_uncertainty", "analysis-limiting", "未识别到重复数、不确定度或重复级记录。", "Agent 无法从统计上区分测量波动与过程偏差。", "优先上传每次重复的原始行及 measurement_repeat_id；至少提供重复数和 SD/SE/CI 的定义。")

    decision_lower = (decision_question or "").lower()
    diagnostic_task = any(token in decision_lower for token in ("偏差", "异常", "失败", "复现", "原因", "机理", "diagnos", "deviation", "why", "error", "uncertainty"))
    if diagnostic_task:
        if "measurement_repeat_id" not in fields:
            add("measurement_repeat_id", "analysis-limiting", "当前任务需要偏差/误差归因，但未识别到重复级 ID。", "只有均值或重复数时，无法区分测量重复性、样品差异和批次差异。", "为每个原始重复增加 measurement_repeat_id，保留未平均的测量值。")
        if "batch_id" not in fields:
            add("batch_id_for_error_budget", "analysis-limiting", "未识别到批次层级。", "批次间变异可能被误归为样品或测量误差。", "增加 batch_id，并说明样品、试样和重复测量的嵌套关系。")
        if not ({"instrument_id", "calibration_id"} & fields):
            add("instrument_and_calibration_provenance", "analysis-limiting", "缺少仪器/校准身份。", "无法判断漂移、设备间差异或标准样问题是否主导偏差。", "提供 instrument_id、calibration_id/日期、标准样结果或校准有效期。")
        if "analysis_version" not in fields:
            add("analysis_version", "optional-enrichment", "未识别到软件、脚本、拟合或分析版本。", "窗口、模型或处理版本变化可能形成数据定义偏差。", "提供分析软件/脚本版本、窗口、拟合模型和删点规则。")
        if not ({"process_actual_value", "measurement_temperature", "measurement_rate"} & fields):
            add("actual_control_and_environment", "optional-enrichment", "只看到结果，未识别到工艺/测试实际控制日志。", "设定值与实测值的差异可能解释协议或过程偏差。", "提供控制器实测值、环境日志或关键条件随时间记录。")
        if not any(token in decision_lower for token in ("只比较", "compare only")):
            add("intermediate_structure_or_property", "optional-enrichment", "未从表头确认能够连接工艺与最终性能的中间结构/性质变量。", "异常传播链可能停留在推测，下一实验难以精准区分机理。", "提供与首要机理相关的结构或中间性质表征；Agent 应先列信息缺口再选择方法。")

    if not ({"process_step", "process_temperature", "process_time", "atmosphere"} & fields):
        add("processing_history", "optional-enrichment", "未识别到工艺历史字段。", "机理推理和文献迁移的针对性会降低。", "提供有顺序的制备/处理步骤及关键参数。")
    return missing


def write_normalized_csv(table: dict[str, Any], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    raw_headers = [column["raw_name"] for column in table["columns"]]
    canonical_headers: list[str] = []
    for column in table["columns"]:
        canonical = column["canonical_name"]
        if canonical and column["mapping_status"] in {"exact", "alias"} and canonical not in canonical_headers:
            canonical_headers.append(canonical)
    path = output_dir / f"{table['table_id']}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=raw_headers + [f"canonical__{name}" for name in canonical_headers])
        writer.writeheader()
        for item in table["rows"]:
            row = dict(item.get("raw", {}))
            for name in canonical_headers:
                value = item.get("canonical", {}).get(name)
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                row[f"canonical__{name}"] = value
            writer.writerow(row)
    return path


def checklist_markdown(packet: dict[str, Any]) -> str:
    readiness = packet["summary"]["readiness"]
    lines = ["# 实验输入检查清单", "", f"状态：**{readiness}**", ""]
    lines.append(f"已登记 {packet['summary']['file_count']} 个文件、{packet['summary']['table_count']} 张表、{packet['summary']['text_count']} 个文本对象。")
    lines.append("")
    for priority, title in (
        ("blocker", "必须补充"),
        ("analysis-limiting", "会限制结论"),
        ("optional-enrichment", "可选增强"),
    ):
        items = [item for item in packet["missing_information"] if item["priority"] == priority]
        lines.extend([f"## {title}", ""])
        if not items:
            lines.append("无。")
        for item in items:
            lines.append(f"- **{item['field']}**：{item['reason']} 影响：{item['impact']} 最小补充：{item['suggested_input']}")
        lines.append("")
    lines.extend(["## 注意", "", "该清单来自确定性文件与表头检查；PDF、图片、自由文本和材料语义仍需要 Agent 进一步阅读。"])
    return "\n".join(lines) + "\n"


def validate_packet(packet: dict[str, Any]) -> list[str]:
    try:
        import jsonschema
    except ImportError:
        return ["jsonschema unavailable; input packet received structural checks only"]
    schema = json.loads(SCHEMA_PATH.read_text(encoding="utf-8"))
    validator = jsonschema.validators.validator_for(schema)(schema)
    errors = []
    for error in sorted(validator.iter_errors(packet), key=lambda item: list(item.path)):
        location = ".".join(str(part) for part in error.absolute_path) or "$"
        errors.append(f"{location}: {error.message}")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Input file or directory")
    parser.add_argument("--output", type=Path, default=Path("intake-output"), help="Output directory")
    parser.add_argument("--decision", default=None, help="Optional R&D decision question")
    parser.add_argument("--material-domain", default=None, help="Optional material domain hint")
    args = parser.parse_args()

    alias_index, canonical_names, alias_spec = load_aliases()
    ambiguous = {normalize_token(key): value for key, value in alias_spec.get("ambiguous_tokens", {}).items()}
    if args.input.resolve() == args.output.resolve():
        parser.error("input and output must be different paths")
    files = iter_files(args.input, args.output)
    args.output.mkdir(parents=True, exist_ok=True)

    packet: dict[str, Any] = {
        "schema_version": "1.0",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "path_base": ".",
        "root_input": relative_artifact_path(args.input, args.output),
        "decision_question": args.decision,
        "material_domain": args.material_domain,
        "received_inputs": [],
        "normalized_tables": [],
        "text_items": [],
        "missing_information": [],
        "summary": {},
    }

    for index, path in enumerate(files, start=1):
        input_id = f"IN-{index:04d}"
        kind, warnings = detect_kind(path)
        record = {
            "id": input_id,
            "path": relative_artifact_path(path, args.output),
            "name": path.name,
            "extension": path.suffix.lower(),
            "size_bytes": path.stat().st_size,
            "sha256": sha256_file(path),
            "detected_kind": kind,
            "parse_status": "manifest-only",
            "warnings": warnings,
        }
        try:
            if kind in {"csv", "tsv"}:
                tables = parse_delimited(path, input_id, alias_index, canonical_names, ambiguous)
                packet["normalized_tables"].extend(tables)
                record["parse_status"] = "parsed"
            elif kind == "xlsx":
                tables, parse_warnings = parse_xlsx(path, input_id, alias_index, canonical_names, ambiguous)
                packet["normalized_tables"].extend(tables)
                record["warnings"].extend(parse_warnings)
                record["parse_status"] = "parsed" if tables else "manifest-only"
            elif kind == "json":
                tables, text_items, parse_warnings = parse_json_file(path, input_id, alias_index, canonical_names, ambiguous)
                packet["normalized_tables"].extend(tables)
                packet["text_items"].extend(text_items)
                record["warnings"].extend(parse_warnings)
                record["parse_status"] = "parsed" if not parse_warnings else "failed"
            elif kind in {"text", "markdown"}:
                item, parse_warnings = parse_text(path, input_id, alias_index, canonical_names, ambiguous)
                packet["text_items"].append(item)
                record["warnings"].extend(parse_warnings)
                record["parse_status"] = "parsed" if item["text"] else "failed"
            else:
                record["parse_status"] = "manifest-only"
        except Exception as exc:  # defensive: preserve manifest even on parser failure
            record["parse_status"] = "failed"
            record["warnings"].append(f"parser failure: {type(exc).__name__}: {exc}")
        packet["received_inputs"].append(record)

    packet["missing_information"] = build_missing(packet, args.decision)
    mapped = 0
    unresolved = 0
    for table in packet["normalized_tables"]:
        for column in table["columns"]:
            if column["mapping_status"] in {"exact", "alias"}:
                mapped += 1
            else:
                unresolved += 1
    blockers = [item for item in packet["missing_information"] if item["priority"] == "blocker"]
    limitations = [item for item in packet["missing_information"] if item["priority"] == "analysis-limiting"]
    readiness = "blocked" if blockers else ("ready-with-caveats" if limitations else "ready")
    packet["summary"] = {
        "file_count": len(packet["received_inputs"]),
        "table_count": len(packet["normalized_tables"]),
        "text_count": len(packet["text_items"]),
        "mapped_field_count": mapped,
        "unresolved_field_count": unresolved,
        "readiness": readiness,
    }

    packet_path = args.output / "intake-packet.json"
    packet_path.write_text(json.dumps(packet, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")
    (args.output / "intake-checklist.md").write_text(checklist_markdown(packet), encoding="utf-8")
    table_dir = args.output / "normalized-tables"
    for table in packet["normalized_tables"]:
        write_normalized_csv(table, table_dir)

    validation_messages = validate_packet(packet)
    hard_errors = [message for message in validation_messages if not message.startswith("jsonschema unavailable")]
    if validation_messages:
        (args.output / "validation-notes.txt").write_text("\n".join(validation_messages) + "\n", encoding="utf-8")
    print(f"WROTE {packet_path}")
    print(f"READINESS {readiness}; files={len(files)}, tables={len(packet['normalized_tables'])}, mapped={mapped}, unresolved={unresolved}")
    if hard_errors:
        for message in hard_errors:
            print(f"ERROR {message}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
